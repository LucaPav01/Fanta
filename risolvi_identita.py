#!/usr/bin/env python3
"""
Risoluzione identità cross-fonte per l'import Fantacalcio.

Fonti coinvolte:
  - fantacalcio-online-csv     : fantacalcio_prezzi.csv (716 righe, quotazioni, no Id)
  - fantacalcio-online-excel   : "Quotazioni Fantacalcio-Online (4).xlsx" (566 righe, anagrafica, no Id)
  - fantacalcio-it-statistiche : "Statistiche_Fantacalcio_Stagione_2026_27 (2).xlsx" (Id presente)
  - fantacalcio-it-quotazioni  : "Quotazioni_Fantacalcio_Stagione_2026_27 (3).xlsx" (Id presente)

Le due fonti fantacalcio.it condividono lo stesso spazio di Id (verificato:
531 Id in comune, zero mismatch nome/squadra sullo stesso Id) => registro
canonico di riferimento con match_method='exact', confidence=1.0 (tier 1:
ID sorgente).

Le due fonti fantacalcio-online non hanno Id: vengono risolte contro il
registro tramite la gerarchia decisa nel progetto:
  1. ID sorgente          -> non applicabile qui (fonti senza Id)
  2. alias confermato      -> non applicabile alla prima esecuzione (nessun
                               alias pregresso da caricare)
  3. name_key(cognome)+team_key esatti
  4. nome univoco (stesso cognome, un solo candidato in tutto il registro,
     squadra diversa - es. giocatore "Estero"/"Serie Minori" nel CSV la cui
     squadra reale è nota solo alla fonte fantacalcio.it)
  5. fuzzy manuale -> mai automatico: tutto ciò che non rientra 1-4 finisce
     in match_review.csv

Nessun merge fuzzy per similarità di stringa viene mai fatto qui.
"""
import csv
import json
import re
import sys
import unicodedata
import uuid
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

from normalizza import make_key, normalize_csv_row

NAMESPACE = uuid.UUID("6e2f6a2e-6b8b-4f6e-9c2a-9a8f5e6d2b10")  # namespace fisso del progetto

ROOT = Path(__file__).resolve().parent
CONFIG = json.loads((ROOT / "app_config.json").read_text(encoding="utf-8"))
SOURCE_FILES = CONFIG["source_files"]
STAT_XLSX = str(ROOT / SOURCE_FILES["fantacalcio_it_statistics"])
QUOT_IT_XLSX = str(ROOT / SOURCE_FILES["fantacalcio_it_quotations"])
QUOT_ONLINE_XLSX = str(ROOT / SOURCE_FILES["fantacalcio_online_excel"])
CSV_PATH = ROOT / SOURCE_FILES["fantacalcio_online_csv"]

DISAMBIGUATOR_RE = re.compile(r"^(.*?)\s+([A-Za-zÀ-ÿ]{1,4})\.$")


def player_uuid(name_key: str, team_key: str) -> str:
    return str(uuid.uuid5(NAMESPACE, f"player:{name_key}:{team_key}"))


def player_uuid_from_source_id(fantacalcio_it_id) -> str:
    """
    Id fantacalcio.it come chiave del player_id canonico per il registro.
    Necessario perché cognome+squadra NON è univoco: due giocatori diversi
    possono condividere entrambi (es. due "Martinez" all'Inter, "Gelli J."
    e "Gelli F." a Frosinone) — usare player_uuid(surname_key, team_key) per
    queste righe li fonderebbe silenziosamente nella stessa entità.
    """
    return str(uuid.uuid5(NAMESPACE, f"player:id:{fantacalcio_it_id}"))


def team_uuid(team_key: str) -> str:
    return str(uuid.uuid5(NAMESPACE, f"team:{team_key}"))


# ------------------------------------------------------------------
# Caricamento fonti
# ------------------------------------------------------------------
def load_xlsx_with_id(path: str, sheet_name: str = "Tutti") -> list[dict]:
    """Fonti fantacalcio.it: prima riga = titolo, seconda = header, colonna 'Id' presente."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    header = None
    out = []
    for row_number, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if header is None:
            if r and r[0] == "Id":
                header = r
            continue
        d = dict(zip(header, r))
        if d.get("Id") is None:
            continue
        d["_row"] = row_number
        out.append(d)
    return out


def load_quot_online(path: str) -> list[dict]:
    """Il foglio ha una riga vuota prima dell'header vero e proprio (colonna 'Nome')."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = None
    out = []
    for row_number, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if header is None:
            if r and r[0] == "Nome":
                header = r
            continue
        d = dict(zip(header, r))
        nome_raw = (d.get("Nome") or "").strip()
        squadra_raw = (d.get("Squadra") or "").strip()
        out.append(
            {
                "row": row_number,
                "nome_raw": nome_raw,
                "squadra_raw": squadra_raw,
                "cognome": extract_cognome_quot_online(nome_raw),
                "team_key": make_key(squadra_raw),
                "raw": d,  # colonne originali integrali, per l'ingestione in player_source_records
            }
        )
    return out


def extract_cognome_quot_online(nome_raw: str) -> str | None:
    """
    Formato sorgente: "COGNOME[ PARTICELLE_MAIUSC] Nome" oppure "COGNOME -"
    (nessun nome fornito). Il cognome è la sequenza di token iniziali
    interamente maiuscoli; il primo token minuscolo/misto (o '-') apre la
    parte nome, che qui non ci serve per il matching.
    """
    tokens = nome_raw.split()
    cognome_tokens = []
    for tok in tokens:
        if tok == "-" or not tok.isupper():
            break
        cognome_tokens.append(tok)
    if not cognome_tokens:
        return None
    return " ".join(cognome_tokens)


def split_registry_surname(nome: str) -> tuple[str, str | None]:
    """
    Formato fantacalcio.it: solo cognome, con disambiguazione "Cognome X."
    (iniziali del nome) quando c'è un omonimo nello stesso pool giocatori.
    Cognomi multi-parola (es. "De Gea", "Van Der Brempt") NON hanno il punto
    finale e vanno quindi presi per intero.
    """
    m = DISAMBIGUATOR_RE.match(nome)
    if m:
        return m.group(1), m.group(2)
    return nome, None


# ------------------------------------------------------------------
# Registro canonico Id-based (fantacalcio-it-statistiche + fantacalcio-it-quotazioni)
# ------------------------------------------------------------------
def build_id_registry(stat_rows: list[dict], quot_rows: list[dict]) -> dict:
    """Un'entità per Id fantacalcio.it. Verificato: nessun mismatch nome/squadra
    sullo stesso Id tra le due fonti, quindi la prima fonte vista vince."""
    registry = {}
    for d in stat_rows + quot_rows:
        rid = d["Id"]
        if rid in registry:
            continue
        base_surname, disamb = split_registry_surname(d["Nome"])
        surname_key = make_key(base_surname)
        team_key = make_key(d["Squadra"])
        registry[rid] = {
            "fantacalcio_it_id": rid,  # ID sorgente: fonte della verità per l'unicità del player_id
            "player_id": player_uuid_from_source_id(rid),
            "nome_raw": d["Nome"],
            "squadra_raw": d["Squadra"],
            "surname_key": surname_key,
            "team_key": team_key,
            "disamb": disamb,
        }
    return registry


def index_registry(registry: dict) -> tuple[dict, dict]:
    by_surname_team = defaultdict(list)
    by_surname = defaultdict(list)
    for entry in registry.values():
        by_surname_team[(entry["surname_key"], entry["team_key"])].append(entry)
        by_surname[entry["surname_key"]].append(entry)
    return by_surname_team, by_surname


def resolve_against_registry(surname_key: str, team_key: str, given_name: str,
                              by_surname_team: dict, by_surname: dict) -> dict:
    """
    Applica la gerarchia (tier 3: name_key+team_key esatti; tier 4: nome
    univoco) contro il registro Id-based. Ritorna un dict con:
      status: 'exact' | 'exact_disambiguated' | 'unique_name_team_mismatch'
              | 'ambiguous' | 'no_counterpart'
      entry / candidates a seconda dello status.
    """
    if not surname_key:
        return {"status": "no_counterpart"}

    same_team = by_surname_team.get((surname_key, team_key), [])
    if len(same_team) == 1:
        return {"status": "exact", "entry": same_team[0]}

    if len(same_team) > 1:
        resolved = _disambiguate(same_team, given_name)
        if resolved is not None:
            return {"status": "exact_disambiguated", "entry": resolved}
        return {"status": "ambiguous", "candidates": same_team, "reason": "omonimi_stessa_squadra"}

    # Nessun candidato con la stessa squadra: prova su tutto il registro.
    all_candidates = by_surname.get(surname_key, [])
    if not all_candidates:
        return {"status": "no_counterpart"}
    if len(all_candidates) == 1:
        return {"status": "unique_name_team_mismatch", "entry": all_candidates[0]}

    resolved = _disambiguate(all_candidates, given_name)
    if resolved is not None:
        return {"status": "exact_disambiguated", "entry": resolved, "team_mismatch": resolved["team_key"] != team_key}
    return {"status": "ambiguous", "candidates": all_candidates, "reason": "omonimi_squadre_diverse"}


def _disambiguate(candidates: list[dict], given_name: str):
    given_name = (given_name or "").strip().upper()
    if not given_name:
        return None
    matches = [c for c in candidates if c["disamb"] and given_name.startswith(c["disamb"].upper())]
    if len(matches) == 1:
        return matches[0]
    return None


# ------------------------------------------------------------------
# Team aliases: verificato che le stringhe squadra sono identiche in tutte
# le fonti (nessuna variante ortografica). "Estero" e "Serie Minori" sono
# categorie del solo CSV (giocatori fuori Serie A), non squadre reali.
# ------------------------------------------------------------------
def build_team_aliases(sources: dict) -> list[dict]:
    seen = set()
    rows = []
    for source_name, team_names in sources.items():
        for raw in sorted(team_names):
            tk = make_key(raw)
            key = (source_name, tk)
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "alias_id": str(uuid.uuid5(NAMESPACE, f"team_alias:{source_name}:{tk}")),
                    "team_id": team_uuid(tk),
                    "alias_raw": raw,
                    "alias_normalized": tk,
                    "source_name": source_name,
                    "canonical_name": raw,
                }
            )
    return rows


# ------------------------------------------------------------------
# Main: esegue la risoluzione su tutte e 4 le fonti e scrive gli output.
# ------------------------------------------------------------------
def main():
    stat_rows = load_xlsx_with_id(STAT_XLSX)
    quot_it_rows = load_xlsx_with_id(QUOT_IT_XLSX)
    registry = build_id_registry(stat_rows, quot_it_rows)
    by_surname_team, by_surname = index_registry(registry)

    with open(CSV_PATH, newline="", encoding="utf-8") as f:
        csv_rows = [normalize_csv_row(row, i + 1) for i, row in enumerate(csv.DictReader(f))]
    quot_online_rows = load_quot_online(QUOT_ONLINE_XLSX)

    player_aliases: list[dict] = []
    review_rows: list[dict] = []
    unresolved_csv = []
    unresolved_qo = []

    def register_alias(source_name, alias_raw, name_key_for_alias, team_key, source_row,
                        canonical_id, canonical_name, canonical_team_key,
                        match_method, match_confidence, note):
        player_aliases.append(
            {
                "alias_id": str(uuid.uuid5(NAMESPACE, f"player_alias:{source_name}:{name_key_for_alias}:{team_key}")),
                "player_id": canonical_id,
                "alias_raw": alias_raw,
                "alias_normalized": name_key_for_alias,
                "source_name": source_name,
                "source_row_number": source_row,
                "canonical_name": canonical_name,
                "canonical_team_key": canonical_team_key,
                "match_method": match_method,
                "match_confidence": match_confidence,
                "note": note,
            }
        )

    def review(source_name, source_row, nome_raw, squadra_raw, ruolo_raw, reason, candidates):
        review_rows.append(
            {
                "source_name": source_name,
                "source_row_number": source_row,
                "nome_raw": nome_raw,
                "squadra_raw": squadra_raw,
                "ruolo_raw": ruolo_raw,
                "issue": reason,
                "candidates": json.dumps(candidates, ensure_ascii=False),
            }
        )

    # --- Alias delle due fonti Id-based verso il proprio registro (tier 1: ID sorgente) ---
    for source_name, rows in (("fantacalcio-it-statistiche", stat_rows), ("fantacalcio-it-quotazioni", quot_it_rows)):
        for d in rows:
            entry = registry[d["Id"]]
            register_alias(source_name, d["Nome"], make_key(d["Nome"]), entry["team_key"], d["_row"],
                            entry["player_id"], entry["nome_raw"], entry["team_key"], "exact", 1.0,
                            f"fantacalcio_it_id={entry['fantacalcio_it_id']}")

    # Il tier 4 ("cognome univoco nel registro") è sicuro solo se il cognome è
    # univoco anche DENTRO la fonte: se la stessa fonte ha più righe con lo
    # stesso cognome (persone diverse, es. i 3 Esposito), non c'è modo di
    # sapere quale delle due corrisponda davvero all'unica voce del registro,
    # quindi tutte finiscono in revisione invece di essere fuse a caso.
    csv_surname_counts = Counter(
        make_key(r["cognome"] or "") for r in csv_rows if not r["name_ambiguous"]
    )
    qo_surname_counts = Counter(
        make_key(r["cognome"]) for r in quot_online_rows if r["cognome"]
    )

    # --- CSV (fantacalcio-online-csv) contro il registro Id-based ---
    for r in csv_rows:
        if r["name_ambiguous"]:
            # già segnalato da normalizza.py: non ha un cognome affidabile
            # da cui partire, resta un caso di revisione manuale a parte.
            unresolved_csv.append(r)
            continue

        surname_key = make_key(r["cognome"] or "")
        outcome = resolve_against_registry(surname_key, r["team_key"], r["nome"] or "",
                                            by_surname_team, by_surname)

        if outcome["status"] in ("exact", "exact_disambiguated"):
            entry = outcome["entry"]
            method = "exact"
            confidence = 1.0
            note = "" if not outcome.get("team_mismatch") else (
                f"nome univoco disambiguato per iniziale; squadra CSV={r['squadra_raw']!r} "
                f"vs registro={entry['squadra_raw']!r}"
            )
            register_alias("fantacalcio-online-csv", r["nome_raw"], r["name_key"], r["team_key"], r["source_row_number"],
                            entry["player_id"], entry["nome_raw"], entry["team_key"], method, confidence, note)
        elif outcome["status"] == "unique_name_team_mismatch" and csv_surname_counts[surname_key] == 1:
            entry = outcome["entry"]
            register_alias(
                "fantacalcio-online-csv", r["nome_raw"], r["name_key"], r["team_key"], r["source_row_number"],
                entry["player_id"], entry["nome_raw"], entry["team_key"], "fuzzy", 0.9,
                f"cognome univoco nel registro, squadra diversa (CSV={r['squadra_raw']!r} "
                f"vs registro={entry['squadra_raw']!r}: probabile prestito/differenza fonte, non un merge di omonimi)",
            )
        elif outcome["status"] == "unique_name_team_mismatch":
            # cognome univoco nel registro ma omonimo dentro il CSV: non è
            # sicuro decidere quale dei due corrisponda alla voce del registro.
            review(
                "fantacalcio-online-csv", r["source_row_number"], r["nome_raw"], r["squadra_raw"], r["ruolo_raw"],
                "cognome univoco nel registro ma omonimo nel CSV: match non sicuro, richiede decisione manuale",
                [{"nome": outcome["entry"]["nome_raw"], "squadra": outcome["entry"]["squadra_raw"],
                  "fantacalcio_it_id": outcome["entry"]["fantacalcio_it_id"]}],
            )
        elif outcome["status"] == "ambiguous":
            review(
                "fantacalcio-online-csv", r["source_row_number"], r["nome_raw"], r["squadra_raw"], r["ruolo_raw"],
                f"omonimi non disambiguabili automaticamente ({outcome['reason']})",
                [{"nome": c["nome_raw"], "squadra": c["squadra_raw"], "fantacalcio_it_id": c["fantacalcio_it_id"]} for c in outcome["candidates"]],
            )
        else:  # no_counterpart
            unresolved_csv.append(r)

    # --- Excel fantacalcio-online contro il registro Id-based ---
    for r in quot_online_rows:
        if not r["cognome"]:
            review("fantacalcio-online-excel", r["row"], r["nome_raw"], r["squadra_raw"], None,
                   "cognome non estraibile dal nome grezzo", [])
            continue

        surname_key = make_key(r["cognome"])
        # Il nome (per la disambiguazione) è tutto ciò che segue il cognome nel nome grezzo.
        given_name = r["nome_raw"][len(r["cognome"]):].strip(" -")
        outcome = resolve_against_registry(surname_key, r["team_key"], given_name, by_surname_team, by_surname)

        if outcome["status"] in ("exact", "exact_disambiguated"):
            entry = outcome["entry"]
            note = "" if not outcome.get("team_mismatch") else (
                f"nome univoco disambiguato per iniziale; squadra Excel={r['squadra_raw']!r} "
                f"vs registro={entry['squadra_raw']!r}"
            )
            register_alias("fantacalcio-online-excel", r["nome_raw"], make_key(r["nome_raw"]), r["team_key"], r["row"],
                            entry["player_id"], entry["nome_raw"], entry["team_key"], "exact", 1.0, note)
        elif outcome["status"] == "unique_name_team_mismatch" and qo_surname_counts[surname_key] == 1:
            entry = outcome["entry"]
            register_alias(
                "fantacalcio-online-excel", r["nome_raw"], make_key(r["nome_raw"]), r["team_key"], r["row"],
                entry["player_id"], entry["nome_raw"], entry["team_key"], "fuzzy", 0.9,
                f"cognome univoco nel registro, squadra diversa (Excel={r['squadra_raw']!r} "
                f"vs registro={entry['squadra_raw']!r})",
            )
        elif outcome["status"] == "unique_name_team_mismatch":
            review(
                "fantacalcio-online-excel", r["row"], r["nome_raw"], r["squadra_raw"], None,
                "cognome univoco nel registro ma omonimo nell'Excel: match non sicuro, richiede decisione manuale",
                [{"nome": outcome["entry"]["nome_raw"], "squadra": outcome["entry"]["squadra_raw"],
                  "fantacalcio_it_id": outcome["entry"]["fantacalcio_it_id"]}],
            )
        elif outcome["status"] == "ambiguous":
            review(
                "fantacalcio-online-excel", r["row"], r["nome_raw"], r["squadra_raw"], None,
                f"omonimi non disambiguabili automaticamente ({outcome['reason']})",
                [{"nome": c["nome_raw"], "squadra": c["squadra_raw"], "fantacalcio_it_id": c["fantacalcio_it_id"]} for c in outcome["candidates"]],
            )
        else:
            unresolved_qo.append(r)

    # --- Incrocio tra le due fonti fantacalcio-online per i residui senza Id ---
    # (giocatori "Estero"/"Serie Minori" o comunque assenti dal registro
    # fantacalcio.it): stessa gerarchia, questa volta CSV<->Excel direttamente.
    qo_by_surname_team = defaultdict(list)
    qo_by_surname = defaultdict(list)
    for r in unresolved_qo:
        sk = make_key(r["cognome"])
        qo_by_surname_team[(sk, r["team_key"])].append(r)
        qo_by_surname[sk].append(r)

    used_qo_ids = set()
    for r in unresolved_csv:
        if r["name_ambiguous"]:
            review("fantacalcio-online-csv", r["source_row_number"], r["nome_raw"], r["squadra_raw"], r["ruolo_raw"],
                   "nome non parsabile automaticamente (split cognome/nome ambiguo)", [])
            continue

        sk = make_key(r["cognome"] or "")
        same_team = qo_by_surname_team.get((sk, r["team_key"]), [])
        all_candidates = qo_by_surname.get(sk, [])
        canonical_id = player_uuid(sk, r["team_key"])

        if len(same_team) == 1 or (len(all_candidates) == 1 and not same_team):
            match = (same_team or all_candidates)[0]
            method, confidence, note = (
                ("exact", 1.0, "match diretto CSV<->Excel (nessun Id fantacalcio.it disponibile)")
                if same_team else
                ("fuzzy", 0.9, f"cognome univoco tra le fonti online, squadra diversa (CSV={r['squadra_raw']!r} "
                               f"vs Excel={match['squadra_raw']!r})")
            )
            used_qo_ids.add(id(match))
            register_alias("fantacalcio-online-csv", r["nome_raw"], r["name_key"], r["team_key"], r["source_row_number"],
                            canonical_id, r["nome_raw"], r["team_key"], method, confidence, note)
            register_alias("fantacalcio-online-excel", match["nome_raw"], make_key(match["nome_raw"]), match["team_key"],
                            match["row"], canonical_id, r["nome_raw"], r["team_key"], method, confidence, note)
        elif len(same_team) > 1 or len(all_candidates) > 1:
            review("fantacalcio-online-csv", r["source_row_number"], r["nome_raw"], r["squadra_raw"], r["ruolo_raw"],
                   "omonimi non disambiguabili tra CSV ed Excel fantacalcio-online",
                   [{"nome": c["nome_raw"], "squadra": c["squadra_raw"]} for c in (same_team or all_candidates)])
        else:
            # Presente solo nel CSV (nessuna fonte lo conferma): entità
            # canonica a sé, nessuna revisione necessaria.
            register_alias("fantacalcio-online-csv", r["nome_raw"], r["name_key"], r["team_key"], r["source_row_number"],
                           canonical_id, r["nome_raw"], r["team_key"], "exact", 1.0,
                           "presente solo in questa fonte, nessun'altra fonte da incrociare")

    for r in unresolved_qo:
        if id(r) in used_qo_ids:
            continue
        sk = make_key(r["cognome"])
        canonical_id = player_uuid(sk, r["team_key"])
        register_alias("fantacalcio-online-excel", r["nome_raw"], make_key(r["nome_raw"]), r["team_key"], r["row"],
                       canonical_id, r["nome_raw"], r["team_key"], "exact", 1.0,
                       "presente solo in questa fonte, nessun'altra fonte da incrociare")

    # --- Team aliases (nessuna variante di stringa rilevata tra le fonti) ---
    team_aliases = build_team_aliases(
        {
            "fantacalcio-online-csv": {r["squadra_raw"] for r in csv_rows},
            "fantacalcio-online-excel": {r["squadra_raw"] for r in quot_online_rows},
            "fantacalcio-it-statistiche": {d["Squadra"] for d in stat_rows},
            "fantacalcio-it-quotazioni": {d["Squadra"] for d in quot_it_rows},
        }
    )

    write_team_aliases(team_aliases)
    write_player_aliases(player_aliases)
    write_match_review(review_rows)

    print(f"Registro Id-based (fantacalcio.it): {len(registry)} giocatori")
    print(f"Team alias generati: {len(team_aliases)}")
    print(f"Player alias generati: {len(player_aliases)}")
    print(f"Righe in revisione manuale: {len(review_rows)}")


def write_team_aliases(rows: list[dict]) -> None:
    with open("team_aliases.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["alias_id", "team_id", "alias_raw", "alias_normalized", "source_name", "canonical_name"])
        writer.writeheader()
        writer.writerows(rows)


def write_player_aliases(rows: list[dict]) -> None:
    fieldnames = [
        "alias_id", "player_id", "alias_raw", "alias_normalized", "source_name",
        "source_row_number", "canonical_name", "canonical_team_key",
        "match_method", "match_confidence", "note",
    ]
    with open("player_aliases.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_match_review(rows: list[dict]) -> None:
    fieldnames = ["source_name", "source_row_number", "nome_raw", "squadra_raw", "ruolo_raw", "issue", "candidates"]
    with open("match_review.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


if __name__ == "__main__":
    main()
